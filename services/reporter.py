import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
import io

from models.schemas import ValidationSummary, ValidationIssue, ChartData, DataStatus
from services.visualizer import generate_static_chart


# Color scheme
COLORS = {
    "header_bg": "2563EB",  # Blue
    "header_fg": "FFFFFF",
    "ready_bg": "10B981",  # Green
    "warning_bg": "F59E0B",  # Yellow/Orange
    "not_ready_bg": "EF4444",  # Red
    "light_gray": "F3F4F6",
    "border": "D1D5DB",
}


def get_status_color(status: DataStatus) -> str:
    """Get background color for status."""
    if status == DataStatus.READY:
        return COLORS["ready_bg"]
    elif status == DataStatus.WARNING:
        return COLORS["warning_bg"]
    return COLORS["not_ready_bg"]


def generate_reports(
    df: pd.DataFrame,
    summary: ValidationSummary,
    issues: List[ValidationIssue],
    charts: List[ChartData],
    output_dir: Path,
    metadata: Dict[str, Any],
    add_watermark: bool = False,
):
    """
    Generate all report files.

    Creates:
    1. cleaned_data.xlsx - The cleaned dataset
    2. validation_report.xlsx - Multi-sheet report with summary, issues, and charts
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # 1. Generate cleaned data file
    generate_cleaned_data_file(df, output_dir, add_watermark)

    # 2. Generate validation report
    generate_validation_report(df, summary, issues, charts, output_dir, metadata, add_watermark)


def generate_cleaned_data_file(df: pd.DataFrame, output_dir: Path, add_watermark: bool = False):
    """Generate the cleaned dataset Excel file."""
    output_path = output_dir / "cleaned_data.xlsx"

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cleaned Data', index=False)

        # Format the header row
        workbook = writer.book
        worksheet = writer.sheets['Cleaned Data']

        header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        header_font = Font(color=COLORS["header_fg"], bold=True)

        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

            # Auto-adjust column width
            col_len = len(str(col))
            if len(df) > 0:
                max_val_len = df[col].astype(str).str.len().max()
                max_val_len = int(max_val_len) if not pd.isna(max_val_len) else 0
            else:
                max_val_len = 0
            max_length = max(col_len, max_val_len)
            worksheet.column_dimensions[cell.column_letter].width = min(max_length + 2, 50)

        # Add watermark for free tier
        if add_watermark:
            watermark_row = len(df) + 3
            worksheet.cell(row=watermark_row, column=1, value="FREE TIER - Data Cleaner Tool")
            worksheet.cell(row=watermark_row, column=1).font = Font(color="999999", italic=True)
            worksheet.cell(row=watermark_row + 1, column=1, value="Upgrade to remove watermark and process unlimited rows")
            worksheet.cell(row=watermark_row + 1, column=1).font = Font(color="999999", italic=True, size=9)


def generate_validation_report(
    df: pd.DataFrame,
    summary: ValidationSummary,
    issues: List[ValidationIssue],
    charts: List[ChartData],
    output_dir: Path,
    metadata: Dict[str, Any],
    add_watermark: bool = False,
):
    """Generate the validation report with multiple sheets."""
    output_path = output_dir / "validation_report.xlsx"

    workbook = Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Sheet 1: Summary
    create_summary_sheet(workbook, summary, metadata, df, add_watermark)

    # Sheet 2: Issue Log
    create_issue_log_sheet(workbook, issues)

    # Sheet 3: Visual Insights
    create_visual_insights_sheet(workbook, charts, output_dir)

    # Sheet 4: Column Details
    create_column_details_sheet(workbook, df, metadata)

    workbook.save(output_path)


def create_summary_sheet(workbook: Workbook, summary: ValidationSummary, metadata: Dict[str, Any], df: pd.DataFrame, add_watermark: bool = False):
    """Create the summary sheet."""
    ws = workbook.create_sheet("Summary")

    # Styles
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_font_white = Font(color=COLORS["header_fg"], bold=True)

    # Title
    ws['A1'] = "Data Quality Report"
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')

    # Watermark for free tier
    if add_watermark:
        ws['A2'] = "FREE TIER - Limited to 100 rows. Upgrade for full access."
        ws['A2'].font = Font(color="FF6600", italic=True, size=10)
        ws.merge_cells('A2:D2')

    # Status badge
    ws['A3'] = "Overall Status:"
    ws['A3'].font = header_font
    ws['B3'] = summary.status.value.upper()
    ws['B3'].fill = PatternFill(start_color=get_status_color(summary.status), end_color=get_status_color(summary.status), fill_type="solid")
    ws['B3'].font = Font(color="FFFFFF", bold=True)
    ws['B3'].alignment = Alignment(horizontal='center')

    ws['A4'] = "Status Reason:"
    ws['B4'] = summary.status_reason
    ws.merge_cells('B4:D4')

    # Metrics table
    ws['A6'] = "Metric"
    ws['B6'] = "Value"
    ws['A6'].fill = header_fill
    ws['B6'].fill = header_fill
    ws['A6'].font = header_font_white
    ws['B6'].font = header_font_white

    metrics = [
        ("Original File", metadata.get("original_filename", "Unknown")),
        ("Total Rows", summary.total_rows),
        ("Total Columns", summary.total_columns),
        ("Missing Values", f"{summary.missing_value_count} ({summary.missing_value_count / (summary.total_rows * summary.total_columns) * 100:.1f}%)" if summary.total_rows > 0 else "0"),
        ("Duplicate Rows", f"{summary.duplicate_row_count} ({summary.duplicate_row_count / summary.total_rows * 100:.1f}%)" if summary.total_rows > 0 else "0"),
        ("Issues Found", summary.issue_count),
    ]

    # Add tier info if truncated
    tier_info = metadata.get("tier_info", {})
    if tier_info.get("truncated"):
        metrics.append(("Note", f"Data truncated from {tier_info.get('original_rows', 'N/A')} to {tier_info.get('processed_rows', 'N/A')} rows (Free tier limit)"))

    for idx, (metric, value) in enumerate(metrics, start=7):
        ws[f'A{idx}'] = metric
        ws[f'B{idx}'] = str(value)
        if idx % 2 == 1:
            ws[f'A{idx}'].fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")
            ws[f'B{idx}'].fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40


def create_issue_log_sheet(workbook: Workbook, issues: List[ValidationIssue]):
    """Create the issue log sheet."""
    ws = workbook.create_sheet("Issue Log")

    # Header styles
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_font = Font(color=COLORS["header_fg"], bold=True)

    # Headers
    headers = ["Issue Type", "Column", "Severity", "Description"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    if not issues:
        ws.cell(row=2, column=1, value="No issues found!")
        ws.merge_cells('A2:D2')
    else:
        for row_idx, issue in enumerate(issues, start=2):
            ws.cell(row=row_idx, column=1, value=issue.issue_type.value)
            ws.cell(row=row_idx, column=2, value=issue.column_name or "N/A")
            ws.cell(row=row_idx, column=3, value=issue.severity.value)
            ws.cell(row=row_idx, column=4, value=issue.description)

            # Color severity
            severity_cell = ws.cell(row=row_idx, column=3)
            if issue.severity.value == "high":
                severity_cell.fill = PatternFill(start_color=COLORS["not_ready_bg"], end_color=COLORS["not_ready_bg"], fill_type="solid")
                severity_cell.font = Font(color="FFFFFF")
            elif issue.severity.value == "medium":
                severity_cell.fill = PatternFill(start_color=COLORS["warning_bg"], end_color=COLORS["warning_bg"], fill_type="solid")
            else:
                severity_cell.fill = PatternFill(start_color=COLORS["ready_bg"], end_color=COLORS["ready_bg"], fill_type="solid")
                severity_cell.font = Font(color="FFFFFF")

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 80


def create_visual_insights_sheet(workbook: Workbook, charts: List[ChartData], output_dir: Path):
    """Create the visual insights sheet with embedded charts."""
    ws = workbook.create_sheet("Visual Insights")

    # Title
    ws['A1'] = "Data Visualizations"
    ws['A1'].font = Font(size=16, bold=True)

    if not charts or len(charts) == 0:
        ws['A3'] = "No charts generated for this dataset."
        return

    # Generate and embed static charts
    row_offset = 3
    chart_height = 20  # Approximate rows per chart

    for idx, chart_data in enumerate(charts):
        try:
            # Generate static chart image
            chart_path = generate_static_chart(chart_data, output_dir)

            # Add chart title
            ws.cell(row=row_offset, column=1, value=chart_data.title)
            ws.cell(row=row_offset, column=1).font = Font(bold=True, size=12)

            # Embed image
            img = XLImage(str(chart_path))
            img.width = 600
            img.height = 350
            ws.add_image(img, f'A{row_offset + 1}')

            row_offset += chart_height

        except Exception as e:
            # If chart generation fails, add text placeholder
            ws.cell(row=row_offset, column=1, value=f"Chart: {chart_data.title}")
            ws.cell(row=row_offset + 1, column=1, value=f"(Chart generation failed: {str(e)})")
            row_offset += 3


def create_column_details_sheet(workbook: Workbook, df: pd.DataFrame, metadata: Dict[str, Any]):
    """Create a sheet with detailed column information."""
    ws = workbook.create_sheet("Column Details")

    # Header styles
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_font = Font(color=COLORS["header_fg"], bold=True)

    # Headers
    headers = ["Column Name", "Original Name", "Data Type", "Non-Null Count", "Null Count", "Unique Values", "Sample Values"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    column_mapping = metadata.get("column_mapping", {})

    for row_idx, col in enumerate(df.columns, start=2):
        ws.cell(row=row_idx, column=1, value=col)
        ws.cell(row=row_idx, column=2, value=column_mapping.get(col, col))
        ws.cell(row=row_idx, column=3, value=str(df[col].dtype))
        ws.cell(row=row_idx, column=4, value=int(df[col].notna().sum()))
        ws.cell(row=row_idx, column=5, value=int(df[col].isna().sum()))
        ws.cell(row=row_idx, column=6, value=int(df[col].nunique()))

        # Sample values
        samples = df[col].dropna().head(3).tolist()
        sample_str = ", ".join([str(s)[:30] for s in samples])
        ws.cell(row=row_idx, column=7, value=sample_str)

        # Alternate row colors
        if row_idx % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=row_idx, column=c).fill = PatternFill(start_color=COLORS["light_gray"], end_color=COLORS["light_gray"], fill_type="solid")

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
